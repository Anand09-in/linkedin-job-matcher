"""
Export routes — Phase 7.6

GET /export/csv    — all matched jobs as CSV download
GET /export/excel  — Excel workbook: 3 sheets
                       1. Matched Jobs  (colour-coded by score)
                       2. Skill Gaps    (frequency table)
                       3. Status Board  (pipeline funnel)
"""
from __future__ import annotations

import csv
import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from loguru import logger
from sqlalchemy import select

from api.dependencies import get_repo
from db.models import Job
from db.repository import AsyncJobRepository

router = APIRouter(prefix="/export", tags=["Export"])


# ── Helpers ───────────────────────────────────────────────────────────────────

_JOB_FIELDS = [
    "id", "title", "company", "location", "date_posted",
    "seniority_level", "employment_type", "remote_policy",
    "experience_years", "salary_range",
    "match_score", "status",
    "matched_skills", "missing_skills",
    "link", "apply_link",
    "scraped_at",
]

def _jobs_to_rows(jobs: list[Job]) -> list[dict]:
    rows = []
    for j in jobs:
        row = {}
        for f in _JOB_FIELDS:
            val = getattr(j, f, None)
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            elif isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d %H:%M")
            elif val is None:
                val = ""
            row[f] = val
        rows.append(row)
    return rows


async def _load_jobs(repo: AsyncJobRepository, has_score: bool = True) -> list[Job]:
    q = select(Job).order_by(Job.match_score.desc().nullslast(), Job.scraped_at.desc())
    if has_score:
        q = q.where(Job.match_score.isnot(None))
    result = await repo._s.execute(q)
    return list(result.scalars().all())


# ── CSV ───────────────────────────────────────────────────────────────────────

@router.get("/csv")
async def export_csv(
    has_score: bool = Query(True, description="Only export matched jobs"),
    min_score: float = Query(0.0, ge=0.0, le=1.0),
    repo: AsyncJobRepository = Depends(get_repo),
):
    """Download all matched jobs as a CSV file."""
    jobs = await _load_jobs(repo, has_score=has_score)
    if min_score > 0:
        jobs = [j for j in jobs if (j.match_score or 0) >= min_score]

    rows = _jobs_to_rows(jobs)

    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=_JOB_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    else:
        output.write("No matched jobs found.")

    output.seek(0)
    filename = f"jobs_export_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    logger.info(f"[Export] CSV: {len(rows)} rows → {filename}")

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Excel ─────────────────────────────────────────────────────────────────────

@router.get("/excel")
async def export_excel(
    has_score: bool = Query(True, description="Only export matched jobs"),
    min_score: float = Query(0.0, ge=0.0, le=1.0),
    repo: AsyncJobRepository = Depends(get_repo),
):
    """Download Excel workbook with 3 sheets: Matched Jobs, Skill Gaps, Status Board."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=501,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    jobs = await _load_jobs(repo, has_score=has_score)
    if min_score > 0:
        jobs = [j for j in jobs if (j.match_score or 0) >= min_score]

    wb = openpyxl.Workbook()

    # ── Sheet 1: Matched Jobs ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Matched Jobs"

    headers = ["Title", "Company", "Location", "Score", "Status",
               "Seniority", "Remote", "Matched Skills", "Missing Skills",
               "Salary", "Experience", "Apply Link", "Scraped"]

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_i, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Score → fill colour
    def score_fill(score: float | None) -> PatternFill:
        if score is None:    return PatternFill("solid", fgColor="F1F5F9")
        if score >= 0.75:    return PatternFill("solid", fgColor="DCFCE7")  # green
        if score >= 0.55:    return PatternFill("solid", fgColor="FEF9C3")  # yellow
        if score >= 0.40:    return PatternFill("solid", fgColor="FED7AA")  # orange
        return PatternFill("solid", fgColor="FEE2E2")                        # red

    for row_i, j in enumerate(jobs, 2):
        score = j.match_score
        fill  = score_fill(score)
        row_data = [
            j.title, j.company, j.location or "",
            f"{score:.0%}" if score is not None else "—",
            j.status,
            j.seniority_level or "", j.remote_policy or "",
            ", ".join(j.matched_skills or []),
            ", ".join(j.missing_skills or []),
            j.salary_range or "", j.experience_years or "",
            j.apply_link or j.link,
            j.scraped_at.strftime("%Y-%m-%d") if j.scraped_at else "",
        ]
        for col_i, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_i, column=col_i, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=(col_i in (8, 9)))

    # Auto-width
    col_widths = [30, 22, 18, 8, 10, 10, 10, 35, 35, 15, 12, 45, 12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Skill Gaps ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Skill Gaps")
    ws2.append(["Skill", "Frequency (jobs)", "% of Jobs", "Priority"])

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    all_missing: list[str] = []
    for j in jobs:
        all_missing.extend(j.missing_skills or [])

    from collections import Counter
    total_jobs = len(jobs)
    counts = Counter(all_missing)
    for skill, count in counts.most_common(30):
        pct = count / total_jobs if total_jobs else 0
        priority = "High" if pct >= 0.6 else ("Medium" if pct >= 0.3 else "Low")
        ws2.append([skill, count, f"{pct:.0%}", priority])

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 10

    # ── Sheet 3: Status Board ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Status Board")
    ws3.append(["Status", "Count", "% of Total"])

    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font

    all_jobs_result = await repo._s.execute(select(Job.status))
    status_counts = Counter(r[0] for r in all_jobs_result.all())
    total_all = sum(status_counts.values())

    status_order = ["new", "saved", "applied", "interview", "offer", "rejected"]
    for s in status_order:
        count = status_counts.get(s, 0)
        pct = count / total_all if total_all else 0
        ws3.append([s.title(), count, f"{pct:.0%}"])

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 12

    # ── Stream workbook ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"jobs_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    logger.info(f"[Export] Excel: {len(jobs)} jobs, {len(counts)} skill gaps → {filename}")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
