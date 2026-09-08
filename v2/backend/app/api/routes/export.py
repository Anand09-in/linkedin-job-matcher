"""
Export routes (Phase 7, ported from v1's api/routes/export.py) —
CSV/Excel downloads of matched jobs, adapted to v2's field set (no v1
`insights`/combined `salary_range` string; salary here is the structured
`salary_benchmark` FR-5 already computed, and experience is
`experience_years_min`, the normalized field FR-2.1 introduced).
"""
from __future__ import annotations

import csv
import io
import uuid
from collections import Counter
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from app.api.dependencies import get_repo
from app.domain.models import Job
from app.domain.repository import Repository

router = APIRouter(prefix="/export", tags=["Export"])

_JOB_FIELDS = [
    "id", "title", "company", "location", "date_posted",
    "seniority_level", "employment_type", "remote_policy",
    "experience_years_min", "salary_benchmark",
    "match_score", "status",
    "matched_skills", "missing_skills",
    "link", "apply_link",
    "scraped_at",
]


def _format_value(val):
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    if isinstance(val, dict):
        # salary_benchmark — render as a compact range string rather than raw JSON
        lo, hi, cur = val.get("min_amount"), val.get("max_amount"), val.get("currency", "")
        if lo is not None or hi is not None:
            return f"{lo or '?'}-{hi or '?'} {cur}".strip()
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    if val is None:
        return ""
    return val


def _jobs_to_rows(jobs: list[Job]) -> list[dict]:
    return [{f: _format_value(getattr(j, f, None)) for f in _JOB_FIELDS} for j in jobs]


async def _load_jobs(repo: Repository, pipeline_id: Optional[uuid.UUID], min_score: float, has_score: bool) -> list[Job]:
    jobs = await repo.list_jobs(
        pipeline_id=pipeline_id, has_score=has_score if has_score else None,
        sort_by="match_score,scraped_at", limit=5000,
    )
    if min_score > 0:
        jobs = [j for j in jobs if (j.match_score or 0) >= min_score]
    return jobs


@router.get("/csv")
async def export_csv(
    pipeline_id: Optional[uuid.UUID] = Query(None),
    has_score: bool = Query(True, description="Only export matched jobs"),
    min_score: float = Query(0.0, ge=0.0, le=1.0),
    repo: Repository = Depends(get_repo),
):
    jobs = await _load_jobs(repo, pipeline_id, min_score, has_score)
    rows = _jobs_to_rows(jobs)

    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=_JOB_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    else:
        output.write("No matched jobs found.")

    filename = f"jobs_export_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/excel")
async def export_excel(
    pipeline_id: Optional[uuid.UUID] = Query(None),
    has_score: bool = Query(True, description="Only export matched jobs"),
    min_score: float = Query(0.0, ge=0.0, le=1.0),
    repo: Repository = Depends(get_repo),
):
    """3 sheets: Matched Jobs (colour-coded by score), Skill Gaps (missing-skill
    frequency across the exported jobs), Status Board (status funnel)."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    jobs = await _load_jobs(repo, pipeline_id, min_score, has_score)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Matched Jobs"
    headers = [
        "Title", "Company", "Location", "Score", "Status", "Seniority", "Remote",
        "Matched Skills", "Missing Skills", "Salary", "Experience", "Apply Link", "Scraped",
    ]
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_i, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    def score_fill(score):
        if score is None:
            return PatternFill("solid", fgColor="F1F5F9")
        if score >= 0.75:
            return PatternFill("solid", fgColor="DCFCE7")
        if score >= 0.55:
            return PatternFill("solid", fgColor="FEF9C3")
        if score >= 0.40:
            return PatternFill("solid", fgColor="FED7AA")
        return PatternFill("solid", fgColor="FEE2E2")

    for row_i, j in enumerate(jobs, 2):
        fill = score_fill(j.match_score)
        row_data = [
            j.title, j.company, j.location or "",
            f"{j.match_score:.0%}" if j.match_score is not None else "—",
            j.status, j.seniority_level or "", j.remote_policy or "",
            ", ".join(j.matched_skills or []), ", ".join(j.missing_skills or []),
            _format_value(j.salary_benchmark), j.experience_years_min or "",
            j.apply_link or j.link, j.scraped_at.strftime("%Y-%m-%d") if j.scraped_at else "",
        ]
        for col_i, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_i, column=col_i, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=(col_i in (8, 9)))

    for i, w in enumerate([30, 22, 18, 8, 10, 10, 10, 35, 35, 15, 12, 45, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Skill Gaps")
    ws2.append(["Skill", "Frequency (jobs)", "% of Jobs", "Priority"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    all_missing: list[str] = []
    for j in jobs:
        all_missing.extend(j.missing_skills or [])
    total_jobs = len(jobs)
    counts = Counter(all_missing)
    for skill, count in counts.most_common(30):
        pct = count / total_jobs if total_jobs else 0
        priority = "High" if pct >= 0.6 else ("Medium" if pct >= 0.3 else "Low")
        ws2.append([skill, count, f"{pct:.0%}", priority])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 10

    ws3 = wb.create_sheet("Status Board")
    ws3.append(["Status", "Count", "% of Total"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    status_counts = Counter(j.status for j in jobs)
    total_all = sum(status_counts.values())
    for s in ["new", "saved", "applied", "interview", "offer", "rejected"]:
        count = status_counts.get(s, 0)
        pct = count / total_all if total_all else 0
        ws3.append([s.title(), count, f"{pct:.0%}"])
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"jobs_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
